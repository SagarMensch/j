import pandas as pd
import openpyxl

input_file = r"C:\Users\sagar\Downloads\Jubilant_Ingrevia_Pricing_with_Justifications.xlsx"
output_file = r"C:\Users\sagar\Downloads\Jubilant_Ingrevia_Pricing_Updated.xlsx"

try:
    # Load workbook using openpyxl directly to preserve formatting as much as possible, or use pandas.
    # Since pandas read/to_excel drops advanced Excel styling, doing it directly via openpyxl is better.
    wb = openpyxl.load_workbook(input_file)
    
    replacements = {
        "Claude Sonnet 4 API": "Fine-Tuned Mistral API",
        "Claude Sonnet 4": "Mistral",
        "Claude": "Mistral Model",
        "Pinecone/Self-hosted": "PostgreSQL (pgvector)",
        "Pinecone/Qdrant": "pgvector inside PostgreSQL",
        "Pinecone Standard": "PostgreSQL (pgvector)",
        "Pinecone": "pgvector",
        "Qdrant": "pgvector",
        "OpenAI/Cohere": "Local Embedding Models (MiniLM)",
        "OpenAI": "Local Models",
        "GPT-4 Turbo": "Proprietary Models (Cost Reference)",
        "GPT-4o": "Standard Proprietary API"
    }
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original_text = cell.value
                    new_text = original_text
                    for old, new in replacements.items():
                        new_text = new_text.replace(old, new)
                    
                    if new_text != original_text:
                        cell.value = new_text

    wb.save(output_file)
    print(f"Successfully saved updated Excel file to: {output_file}")
except Exception as e:
    print(f"Failed to update Excel file: {e}")
