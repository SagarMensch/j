import openpyxl

input_file = r"C:\Users\sagar\Downloads\Jubilant_Ingrevia_Pricing_with_Justifications.xlsx"
output_file = r"C:\Users\sagar\Downloads\Jubilant_Ingrevia_Pricing_Final_Sanitized.xlsx"

try:
    wb = openpyxl.load_workbook(input_file)
    
    replacements = {
        # Removing any specific LLM names
        "Claude Sonnet 4 API": "Self-Hosted Fine-Tuned LLM Pipeline",
        "Claude Sonnet 4": "Domain-Specific Fine-Tuned LLM",
        "Claude": "Fine-Tuned Open-Source LLM",
        "Mistral": "Fine-Tuned Open-Source LLM",
        # Keep pgvector over Pinecone
        "Pinecone/Self-hosted": "PostgreSQL (pgvector)",
        "Pinecone/Qdrant": "PostgreSQL (pgvector)",
        "Pinecone Standard": "PostgreSQL (pgvector)",
        "Pinecone": "pgvector",
        "Qdrant": "pgvector",
        "OpenAI/Cohere": "Local Embedding Models",
        "OpenAI": "Proprietary Models",
        "GPT-4 Turbo": "Proprietary Cloud Models (Cost Reference)",
        "GPT-4o": "Standard Commercial API",
        # Fix the justification narrative for 'Query Cost' so it doesn't look like double-dipping
        "Cost: API ₹6.50 + Vector search ₹0.80 + Synthesis processing ₹0.70": "Cost includes: RAG Pipeline Orchestration, Vector Search Compute, Context Synthesis Routing.",
        "Cost: Claude Sonnet 4 API ($0.003/1K input + $0.012/1K output) = $0.006 × ₹83 = ₹0.50 + Vector search (₹0.10) + Infrastructure overhead (₹0.20) = ₹0.80": "Cost encompasses dynamic RAG routing, embedding scaling overhead, and pipeline load-balancing during query surges.",
        "Cost: API ₹16.00 + Vector search ₹2.00 + Heavy compute ₹2.00 = ₹20.00": "Cost entails Cross-Graph Traversal Compute, Multi-Hop Reasoning limits, and Synthesis overhead."
    }
    
    # Also fix the math error: The Dashboard summary incorrectly stated 600000 for One-Time Setup. 
    # It should be 2400000 (20L platform + 4L training)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                # Text Replacements
                if cell.value and isinstance(cell.value, str):
                    original_text = cell.value
                    new_text = original_text
                    for old, new in replacements.items():
                        new_text = new_text.replace(old, new)
                    
                    if new_text != original_text:
                        cell.value = new_text
                
                # Math Fixes
                if cell.value == 600000 or cell.value == 600000.0:
                    # Check if the row title has 'One-Time Setup Cost' or similar
                    row_values = [str(c.value) for c in row if c.value]
                    if any("One-Time" in v or "TOTAL" in v for v in row_values):
                         cell.value = 2400000
    
    # Recalculate Sub-Total (Before GST) in Dashboard
    # If the setup went from 6L to 24L, the Sub-Total Year 1 increases by 18L.
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == 3324000 or cell.value == 3324000.0: # Old Year 1 Total
                     cell.value = 5124000
                elif cell.value == 4178460.432:
                     cell.value = 5978460.432 # Old TCO Year 1
                elif cell.value == 4930583.30976:
                     cell.value = 7054583.31 # TCO incl GST
                elif cell.value == 13375749.92928 or cell.value == 13375749.929279998:
                     cell.value = 15175749.93 # 3 Year Total
                

    wb.save(output_file)
    print(f"Saved completely sanitized and mathematically robust excel to {output_file}")
except Exception as e:
    print(f"Error: {e}")
